from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime
from typing import Any


class DocumentService:
    def extract_text(self, buffer: bytes, filename: str) -> str:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext == "pdf":
            return self.clean_text(self._extract_pdf(buffer))
        if ext == "docx":
            return self.clean_text(self._extract_docx(buffer))
        if ext in {"txt", "text"}:
            return self.clean_text(buffer.decode("utf-8", errors="ignore"))
        if ext in {"xlsx", "xls"}:
            return self.clean_text(self._extract_excel(buffer))

        raise RuntimeError(f"Unsupported file type: {ext}")

    def _extract_pdf(self, buffer: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Missing dependency pypdf. Run: pip install -r requirements.txt") from exc

        reader = PdfReader(io.BytesIO(buffer))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)

    def _extract_docx(self, buffer: bytes) -> str:
        try:
            import mammoth
        except ImportError as exc:
            raise RuntimeError("Missing dependency mammoth. Run: pip install -r requirements.txt") from exc

        result = mammoth.extract_raw_text(io.BytesIO(buffer))
        return result.value or ""

    def _extract_excel(self, buffer: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Missing dependency openpyxl. Run: pip install -r requirements.txt") from exc

        workbook = load_workbook(io.BytesIO(buffer), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if values:
                    lines.append(", ".join(values))
        return "\n".join(lines)

    def clean_text(self, text: str) -> str:
        if not text:
            return ""

        cleaned = text.replace("\r\n", "\n").replace("\t", " ")
        cleaned = re.sub(r"[ ]{2,}", " ", cleaned)
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
        return cleaned.strip()

    def parse_metadata_from_path(self, file_path: str, file_content: str | None = None) -> dict[str, Any]:
        normalized = file_path.replace("\\", "/")
        parts = normalized.split("/")
        filename = parts[-1] if parts else normalized

        folder = "syllabus"
        for item in parts:
            if item in {"syllabus", "curriculum", "regulation", "career", "career description", "career_description", "career-description"}:
                folder = item
                break

        if folder == "curriculum":
            return self._parse_curriculum_metadata(filename, file_content)
        if folder in {"career", "career description", "career_description", "career-description"}:
            return self._parse_career_metadata(filename, file_content)
        if folder == "regulation":
            return self._parse_regulation_metadata(filename, file_content)
        return self._parse_syllabus_metadata(filename, file_content)

    def _parse_syllabus_metadata(self, filename: str, content: str | None) -> dict[str, Any]:
        clean_name = filename.replace(".docx", "").strip()
        parts = clean_name.split("_", 1)
        subject_name = parts[0] if parts else clean_name
        subject_code = parts[1] if len(parts) > 1 else self._extract_subject_code(clean_name)

        return {
            "doc_type": "syllabus",
            "document_type": "syllabus",
            "subject_name": subject_name,
            "subject_code": subject_code,
            "course_code": subject_code,
            "major": self._detect_major(clean_name),
            "credits": self._extract_credits(content),
            "faculty": self._detect_faculty(clean_name),
            "level": "undergraduate",
            "language": "vi",
            "academic_year": self._extract_academic_year(content) or "2024-2025",
            "source_file": filename,
        }

    def _parse_curriculum_metadata(self, filename: str, content: str | None) -> dict[str, Any]:
        clean_name = filename.replace(".docx", "").lower()
        major = self._detect_major_full(clean_name)
        return {
            "doc_type": "curriculum",
            "document_type": "curriculum",
            "program_name": self._extract_program_name(clean_name),
            "major": major,
            "major_code": self._get_major_code(major),
            "degree": "Bachelor",
            "total_credits": self._extract_total_credits(content) or 130,
            "training_duration": "4 years",
            "admission_from_year": self._extract_admission_year(content) or 2024,
            "issuing_decision": self._extract_decision_number(content) or "N/A",
            "issuing_date": self._extract_issuing_date(content) or "N/A",
            "managing_unit": self._detect_faculty(clean_name),
            "language": "vi",
            "source_file": filename,
        }

    def _parse_regulation_metadata(self, filename: str, content: str | None) -> dict[str, Any]:
        clean_name = filename.replace(".docx", "").lower()
        is_admission = "tuyen sinh" in self._normalize_for_search(clean_name) or "de an" in self._normalize_for_search(clean_name)

        if is_admission:
            return {
                "doc_type": "regulation",
                "document_type": "regulation",
                "regulation_type": "admission_policy",
                "admission_year": self._extract_year(clean_name) or datetime.utcnow().year,
                "education_level": "undergraduate",
                "institution": "National Economics University",
                "applicable_major": "all",
                "issuing_body": "Admission Council",
                "language": "vi",
                "source_file": filename,
            }

        return {
            "doc_type": "regulation",
            "document_type": "regulation",
            "regulation_type": "student_assessment",
            "decision_number": self._extract_decision_number(content),
            "issued_year": self._extract_year(clean_name),
            "issuing_body": "National Economics University",
            "applicable_object": "students",
            "effective_status": "expired" if self._is_expired(content) else "active",
            "language": "vi",
            "source_file": filename,
        }

    def _parse_career_metadata(self, filename: str, content: str | None) -> dict[str, Any]:
        clean_name = filename.rsplit(".", 1)[0].strip()
        career_name_en = re.sub(r"[_\-]+", " ", clean_name).strip()
        section = self._detect_career_section(clean_name, content)

        return {
            "doc_type": "career",
            "document_type": "career",
            "career_name_en": career_name_en,
            "section": section,
            "language": "vi",
            "source_file": filename,
        }

    def _extract_subject_code(self, text: str) -> str:
        match = re.search(r"[A-Z]{2,4}\d{3,4}", text, flags=re.IGNORECASE)
        return match.group(0).upper() if match else "UNKNOWN"

    def _detect_major(self, text: str) -> str:
        lower = self._normalize_for_search(text)
        if "cong nghe thong tin" in lower or "cntt" in lower:
            return "Cong nghe thong tin"
        if "khoa hoc may tinh" in lower or "khmt" in lower:
            return "Khoa hoc may tinh"
        if "kinh te" in lower:
            return "Kinh te"
        if "marketing" in lower:
            return "Marketing"
        return "Khac"

    def _detect_major_full(self, text: str) -> str:
        lower = self._normalize_for_search(text)
        if "cong nghe thong tin" in lower or "cntt" in lower:
            return "Cong nghe thong tin"
        if "khoa hoc may tinh" in lower or "khmt" in lower:
            return "Khoa hoc may tinh"
        if "kinh te" in lower:
            return "Kinh te"
        if "marketing" in lower:
            return "Marketing"
        if "ke toan" in lower:
            return "Ke toan"
        return "Da nganh"

    def _get_major_code(self, major: str) -> str:
        codes = {
            "Cong nghe thong tin": "7480201",
            "Khoa hoc may tinh": "7480101",
            "Kinh te": "7310106",
            "Marketing": "7340115",
            "Ke toan": "7340301",
        }
        return codes.get(major, "7000000")

    def _detect_faculty(self, text: str) -> str:
        lower = self._normalize_for_search(text)
        if "cntt" in lower or "cong nghe thong tin" in lower:
            return "Institute of Information Technology"
        if "kinh te" in lower:
            return "Faculty of Economics"
        if "marketing" in lower:
            return "Faculty of Marketing"
        return "National Economics University"

    def _extract_credits(self, content: str | None) -> int:
        if not content:
            return 3

        normalized = self._normalize_for_search(content)
        patterns = [
            r"so\s*tin\s*chi\s*[:\-]?\s*(\d{1,2})",
            r"tin\s*chi\s*[:\-]?\s*(\d{1,2})",
            r"credits?\s*[:\-]?\s*(\d{1,2})",
            r"(\d{1,2})\s*(tin\s*chi|tc|credits?)\b",
        ]

        values: list[int] = []
        for pattern in patterns:
            match = re.search(pattern, normalized, flags=re.IGNORECASE)
            if match:
                value = int(match.group(1))
                if value > 0:
                    values.append(value)

        return max(values) if values else 3

    def _extract_total_credits(self, content: str | None) -> int | None:
        if not content:
            return None
        match = re.search(r"tong\s*so.*?(\d{2,3})\s*(?:tin\s*chi|credits)", self._normalize_for_search(content))
        return int(match.group(1)) if match else None

    def _extract_academic_year(self, content: str | None) -> str | None:
        if not content:
            return None
        match = re.search(r"(?:nam\s*hoc\s*)?(\d{4})-?(\d{4})?", self._normalize_for_search(content), flags=re.IGNORECASE)
        if not match:
            return None
        return f"{match.group(1)}-{match.group(2)}" if match.group(2) else match.group(1)

    def _extract_admission_year(self, content: str | None) -> int | None:
        if not content:
            return None
        match = re.search(r"khoa\s*(\d{4})", self._normalize_for_search(content), flags=re.IGNORECASE)
        return int(match.group(1)) if match else None

    def _extract_decision_number(self, content: str | None) -> str | None:
        if not content:
            return None
        match = re.search(r"(?:so\s*)?(\d+\/[A-Z]{2}-[A-ZD]+)", content, flags=re.IGNORECASE)
        return match.group(1) if match else None

    def _extract_issuing_date(self, content: str | None) -> str | None:
        if not content:
            return None
        match = re.search(r"(\d{1,2}[\/-]\d{1,2}[\/-]\d{4})", content)
        return match.group(1) if match else None

    def _extract_year(self, text: str | None) -> int | None:
        if not text:
            return None
        match = re.search(r"\b(20\d{2})\b", text)
        return int(match.group(1)) if match else None

    def _extract_program_name(self, text: str) -> str:
        normalized = self._normalize_for_search(text)
        if "cong nghe thong tin" in normalized:
            return "Bachelor of Information Technology"
        if "khoa hoc may tinh" in normalized:
            return "Bachelor of Computer Science"
        return "Training Program"

    def _detect_career_section(self, file_name: str, content: str | None) -> str:
        haystack = self._normalize_for_search(f"{file_name}\n{content or ''}")
        if "ky nang" in haystack or "skill" in haystack:
            return "skills"
        if "nhiem vu" in haystack or "trach nhiem" in haystack or "responsibilit" in haystack:
            return "responsibilities"
        if "yeu cau" in haystack or "requirement" in haystack:
            return "requirements"
        return "overview"

    def _is_expired(self, content: str | None) -> bool:
        if not content:
            return False

        normalized = self._normalize_for_search(content)
        if "het hieu luc" in normalized or "da thay the" in normalized:
            return True

        year = self._extract_year(normalized)
        return bool(year and year < 2020)

    def _normalize_for_search(self, text: str) -> str:
        normalized = unicodedata.normalize("NFD", str(text or "").lower())
        normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.strip()

    def chunk_text(self, text: str, chunk_size: int = 700, overlap: int = 200) -> list[str]:
        if not text:
            return []

        chunks: list[str] = []
        start = 0
        while start < len(text):
            end = min(start + chunk_size, len(text))
            section = text[start:end].strip()
            if len(section) > 50:
                chunks.append(section)
            start += max(1, chunk_size - overlap)
        return chunks

